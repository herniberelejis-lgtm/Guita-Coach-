"""Investment CSV/XLSX parser for Argentine brokers: Cocos Capital, Invertir Online, Bull Market.

Each broker exports transactions in a different format. This module:
1. Auto-detects the broker format by inspecting headers
2. Parses the file according to broker-specific format
3. Returns standardized transaction dicts with keys:
   - date (YYYY-MM-DD format)
   - tx_type ("buy" or "sell")
   - ticker (uppercase)
   - quantity (float)
   - price (float)
   - broker (string)
   - csv_reference (for audit trail)

Handles edge cases:
- Empty rows and malformed data (skipped)
- Multiple date formats (YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY)
- European and US number formats (1.234,56 vs 1,234.56)
- Non-ASCII characters
- UTF-8 and Latin-1 encodings
- Excel files (.xlsx) with openpyxl
"""
import csv
import io
import re
from typing import Optional, Tuple

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


_DATE_PATTERNS = (
    (re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})"), lambda m: f"{m[1]}-{int(m[2]):02d}-{int(m[3]):02d}"),
    (re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})"), lambda m: f"{m[3]}-{int(m[2]):02d}-{int(m[1]):02d}"),
    (re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{4})"), lambda m: f"{m[3]}-{int(m[2]):02d}-{int(m[1]):02d}"),
)


def _parse_date(raw: str) -> Optional[str]:
    """Parse date in multiple formats to YYYY-MM-DD.

    Supports: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, with or without zero-padding
    (e.g. "5/3/2026" as well as "05/03/2026"). Returns None if date cannot be parsed.
    """
    raw = (raw or "").strip()
    for pattern, build in _DATE_PATTERNS:
        m = pattern.match(raw)
        if m:
            return build(m)
    return None


def _parse_amount(raw: str) -> Optional[float]:
    """Parse numeric amount handling European and US formats.

    Supports:
    - '-1.234,56' (European: thousands separator=dot, decimal=comma)
    - '1,234.56' (US: thousands separator=comma, decimal=dot)
    - '22,1225' (European with 4 decimals)
    - '$ 500' (with currency symbol)
    - '(300,00)' (parentheses for negative)

    Returns None if amount cannot be parsed.
    """
    raw = (raw or "").strip().replace("$", "").replace(" ", "")
    if not raw:
        return None

    negative = raw.startswith("-") or (raw.startswith("(") and raw.endswith(")"))
    raw = raw.strip("()-+")

    # Detect if we have both comma and dot
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            # European: 1.234,56 (comma is rightmost = decimal)
            raw = raw.replace(".", "").replace(",", ".")
        else:
            # US: 1,234.56 (dot is rightmost = decimal)
            raw = raw.replace(",", "")
    elif "," in raw:
        # Only comma: could be European (1,23) or US thousands (1,000)
        # If comma leaves more than 2 digits after, it's decimal (e.g. 22,1225)
        # If comma leaves 1-2 digits after, it's decimal (e.g. 1,23)
        head, _, tail = raw.rpartition(",")
        if head and all(c.isdigit() for c in head):
            # It's a valid number, comma is decimal
            raw = f"{head.replace('.', '')}.{tail}"
        else:
            raw = raw.replace(",", "")  # US thousands or invalid
    elif "." in raw:
        # Only dot: could be thousands separator or decimal point
        # If dot leaves more than 2 digits, it's thousands separator
        # If dot leaves 1-2 digits, it's decimal
        head, _, tail = raw.rpartition(".")
        if len(tail) <= 2 and head and all(c.isdigit() for c in head):
            # Decimal point (e.g., "123.45" or "123.5")
            raw = f"{head}.{tail}"
        else:
            # Thousands separator or invalid
            raw = raw.replace(".", "")

    try:
        value = float(raw)
    except ValueError:
        return None

    return -value if negative else value


def _read_text(content: bytes) -> Optional[str]:
    """Decode bytes to text trying multiple encodings."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return None


def _detect_delimiter(text: str) -> str:
    """Detect CSV delimiter by counting ; vs , in first few lines."""
    first_lines = "\n".join(text.split("\n")[:5])
    return ";" if first_lines.count(";") > first_lines.count(",") else ","


def _get_headers_lower(row: list[str]) -> list[str]:
    """Return list of lowercase, stripped header names."""
    return [cell.lower().strip() for cell in row]


def _is_cocos_account_statement(headers: list[str]) -> bool:
    """Detect Cocos Capital's real account-statement export.

    Columns: nroTicket;nroComprobante;fechaEjecucion;fechaLiquidacion;tipoOperacion;
    instrumento;moneda;mercado;cantidad;precio;montoBruto;comision;...;total
    """
    return "nroticket" in headers and "tipooperacion" in headers and "instrumento" in headers


def detect_broker(csv_bytes: bytes) -> Optional[str]:
    """Auto-detect broker format from CSV headers.

    Returns one of: "cocos_capital", "invertir_online", "bull_market", or None.

    Detection strategy:
    - Cocos Capital (account statement): has 'nroticket' AND 'tipooperacion' AND 'instrumento'
    - Cocos Capital (simple format): has 'especie' AND 'comision' in headers
    - Invertir Online: has 'isin' AND 'liquidacion' in headers
    - Bull Market: has 'simbolo' OR ('instrumento' AND 'tipooperacion') in headers
    """
    text = _read_text(csv_bytes)
    if not text:
        return None

    delimiter = _detect_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))

    if not rows:
        return None

    headers = _get_headers_lower(rows[0])

    if _is_cocos_account_statement(headers):
        return "cocos_capital"

    # Check for Cocos Capital (has both especie and comision)
    if "especie" in headers and "comision" in headers:
        return "cocos_capital"

    # Check for Invertir Online (has both isin and liquidacion)
    if "isin" in headers and "liquidacion" in headers:
        return "invertir_online"

    # Check for Bull Market (has simbolo OR instrumento+tipooperacion)
    if "simbolo" in headers:
        return "bull_market"
    if "instrumento" in headers and "tipooperacion" in headers:
        return "bull_market"

    return None


def _cocos_account_statement_tx_type(tipo_operacion: str) -> Optional[str]:
    """Map a Cocos tipoOperacion value to "buy"/"sell", or None if it's not a trade.

    Covers plain Compra/Venta and FCI movements (Liquidacion Suscripcion Fci = buy,
    Liquidacion Rescate Fci = sell). "Operatoria dolar MEP" pairs (buy the bond in
    one currency leg, sell it in the other) are a currency-conversion mechanism, not
    a real position in the bond, so they're excluded like cash movements (Recibo De
    Cobro, Orden De Pago, Dividendos en especie) — otherwise the bond shows up as a
    large phantom holding once the two legs' mismatched per-currency prices get
    averaged together.
    """
    t = tipo_operacion.lower()
    if "dolar mep" in t or "dólar mep" in t:
        return None
    if "venta" in t or "rescate" in t:
        return "sell"
    if "compra" in t or "suscripcion" in t:
        return "buy"
    return None


def _extract_ticker_from_instrumento(instrumento: str) -> str:
    """Extract the ticker from a Cocos 'instrumento' description like
    'CEDEAR DE MICROSOFT CORP. (MSFT)' -> 'MSFT'."""
    raw = instrumento.strip()
    if "(" in raw and ")" in raw:
        raw = raw[raw.rfind("(") + 1:raw.rfind(")")]
    return raw.upper()


def _parse_cocos_account_statement(rows: list[list[str]]) -> list[dict]:
    """Parse Cocos Capital's real account-statement export (rows already split).

    Columns: nroTicket;nroComprobante;fechaEjecucion;fechaLiquidacion;tipoOperacion;
    instrumento;moneda;mercado;cantidad;precio;montoBruto;comision;...;total

    tipoOperacion drives buy/sell direction (quantity sign in the source file is
    not reliable for FCI redemptions, so quantity is always normalized to abs()).
    Cash movements (deposits, withdrawals, in-kind dividends) are skipped.
    """
    if not rows:
        return []

    headers = _get_headers_lower(rows[0])
    col_indices = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col_indices["fecha"] = i
        elif "tipooperacion" in h:
            col_indices["tipooperacion"] = i
        elif "instrumento" in h:
            col_indices["instrumento"] = i
        elif "cantidad" in h:
            col_indices["cantidad"] = i
        elif "precio" in h:
            col_indices["precio"] = i

    required = {"fecha", "tipooperacion", "instrumento", "cantidad", "precio"}
    if not required.issubset(col_indices.keys()):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col_indices.values()):
            continue

        tx_type = _cocos_account_statement_tx_type(str(row[col_indices["tipooperacion"]]))
        if tx_type is None:
            continue

        date = _parse_date(str(row[col_indices["fecha"]]))
        if not date:
            continue

        ticker = _extract_ticker_from_instrumento(str(row[col_indices["instrumento"]]))
        if not ticker:
            continue

        quantity = _parse_amount(str(row[col_indices["cantidad"]]))
        if quantity is None or quantity == 0:
            continue
        quantity = abs(quantity)

        price = _parse_amount(str(row[col_indices["precio"]]))
        if price is None or price <= 0:
            continue

        csv_reference = f"cocos_{date}_{ticker}_{quantity}"

        items.append({
            "date": date,
            "tx_type": tx_type,
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": "cocos_capital",
            "csv_reference": csv_reference,
        })

    return items


def _parse_cocos_simple(rows: list[list[str]]) -> list[dict]:
    """Parse Cocos Capital's simple format: Fecha, Tipo, Especie, Cantidad, Precio, Comision, Total."""
    if not rows:
        return []

    headers = _get_headers_lower(rows[0])

    col_indices = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col_indices["fecha"] = i
        elif "tipo" in h:
            col_indices["tipo"] = i
        elif "especie" in h:
            col_indices["especie"] = i
        elif "cantidad" in h:
            col_indices["cantidad"] = i
        elif "precio" in h:
            col_indices["precio"] = i

    required = {"fecha", "tipo", "especie", "cantidad", "precio"}
    if not required.issubset(col_indices.keys()):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col_indices.values()):
            continue

        date = _parse_date(row[col_indices["fecha"]])
        if not date:
            continue

        tx_type_raw = row[col_indices["tipo"]].strip()
        tx_type = "sell" if "venta" in tx_type_raw.lower() else "buy"

        ticker = row[col_indices["especie"]].strip().upper()
        if not ticker:
            continue

        quantity = _parse_amount(row[col_indices["cantidad"]])
        if quantity is None or quantity <= 0:
            continue

        price = _parse_amount(row[col_indices["precio"]])
        if price is None or price <= 0:
            continue

        csv_reference = f"cocos_{date}_{ticker}_{quantity}"

        items.append({
            "date": date,
            "tx_type": tx_type,
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": "cocos_capital",
            "csv_reference": csv_reference,
        })

    return items


def parse_cocos_csv(csv_bytes: bytes) -> list[dict]:
    """Parse Cocos Capital CSV, auto-detecting between the real account-statement
    export and the simple Fecha/Tipo/Especie/Cantidad/Precio format."""
    text = _read_text(csv_bytes)
    if not text:
        return []

    delimiter = _detect_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))

    if not rows:
        return []

    headers = _get_headers_lower(rows[0])
    if _is_cocos_account_statement(headers):
        return _parse_cocos_account_statement(rows)
    return _parse_cocos_simple(rows)


def parse_invertir_online_csv(csv_bytes: bytes) -> list[dict]:
    """Parse Invertir Online CSV format.

    Expected columns: Fecha, ISIN, Especie, Cantidad, Precio, Liquidacion, Total
    Note: Invertir Online doesn't export transaction type, so all are treated as buy.
    """
    text = _read_text(csv_bytes)
    if not text:
        return []

    delimiter = _detect_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))

    if not rows:
        return []

    headers = _get_headers_lower(rows[0])

    # Find column indices
    col_indices = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col_indices["fecha"] = i
        elif "isin" in h:
            col_indices["isin"] = i
        elif "especie" in h:
            col_indices["especie"] = i
        elif "cantidad" in h:
            col_indices["cantidad"] = i
        elif "precio" in h:
            col_indices["precio"] = i

    # Validate we have minimum required columns
    required = {"fecha", "especie", "cantidad", "precio"}
    if not required.issubset(col_indices.keys()):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col_indices.values()):
            continue

        date = _parse_date(row[col_indices["fecha"]])
        if not date:
            continue

        ticker = row[col_indices["especie"]].strip().upper()
        if not ticker:
            continue

        quantity = _parse_amount(row[col_indices["cantidad"]])
        if quantity is None or quantity <= 0:
            continue

        price = _parse_amount(row[col_indices["precio"]])
        if price is None or price <= 0:
            continue

        isin = row[col_indices["isin"]].strip() if "isin" in col_indices else ""
        csv_reference = f"io_{isin}_{date}_{ticker}" if isin else f"io_{date}_{ticker}"

        items.append({
            "date": date,
            "tx_type": "buy",  # Invertir Online doesn't export transaction type
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": "invertir_online",
            "csv_reference": csv_reference,
        })

    return items


def parse_bull_market_csv(csv_bytes: bytes) -> list[dict]:
    """Parse Bull Market CSV format.

    Supports two formats:
    1. Simple: Fecha, Simbolo, Cantidad, Precio
    2. Full: fechaEjecucion, instrumento, tipoOperacion, cantidad, precio

    Note: Old format treats all as buy. New format detects Compra/Venta.
    """
    text = _read_text(csv_bytes)
    if not text:
        return []

    delimiter = _detect_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))

    if not rows:
        return []

    headers = _get_headers_lower(rows[0])

    # Find column indices
    col_indices = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col_indices["fecha"] = i
        elif "simbolo" in h:
            col_indices["simbolo"] = i
        elif "instrumento" in h:
            col_indices["instrumento"] = i
        elif "tipooperacion" in h:
            col_indices["tipooperacion"] = i
        elif "cantidad" in h:
            col_indices["cantidad"] = i
        elif "precio" in h:
            col_indices["precio"] = i

    # Validate we have minimum required columns
    has_fecha = "fecha" in col_indices
    has_ticker = "simbolo" in col_indices or "instrumento" in col_indices
    has_qty = "cantidad" in col_indices
    has_price = "precio" in col_indices

    if not (has_fecha and has_ticker and has_qty and has_price):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col_indices.values()):
            continue

        date = _parse_date(row[col_indices["fecha"]])
        if not date:
            continue

        # Extract ticker from either simbolo or instrumento
        ticker_raw = ""
        if "simbolo" in col_indices:
            ticker_raw = row[col_indices["simbolo"]].strip()
        elif "instrumento" in col_indices:
            ticker_raw = row[col_indices["instrumento"]].strip()
            # Extract ticker from "COMPANY NAME (TICKER)" format
            if "(" in ticker_raw and ")" in ticker_raw:
                ticker_raw = ticker_raw[ticker_raw.rfind("(") + 1:ticker_raw.rfind(")")]

        ticker = ticker_raw.upper()
        if not ticker:
            continue

        quantity_raw = row[col_indices["cantidad"]]
        quantity = _parse_amount(quantity_raw)
        if quantity is None or quantity == 0:
            continue

        price_raw = row[col_indices["precio"]]
        price = _parse_amount(price_raw)
        if price is None or price <= 0:
            continue

        # Detect transaction type from tipoOperacion if available
        tx_type = "buy"
        if "tipooperacion" in col_indices:
            tx_type_raw = row[col_indices["tipooperacion"]].strip().lower()
            if "venta" in tx_type_raw:
                tx_type = "sell"
                quantity = abs(quantity)  # Ensure quantity is positive for sells
            elif "compra" in tx_type_raw:
                tx_type = "buy"
                quantity = abs(quantity)

        csv_reference = f"bull_{date}_{ticker}_{quantity}"

        items.append({
            "date": date,
            "tx_type": tx_type,
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": "bull_market",
            "csv_reference": csv_reference,
        })

    return items


def parse_csv(csv_bytes: bytes) -> Tuple[Optional[str], list[dict]]:
    """Auto-detect broker and parse CSV.

    Returns tuple of (broker_name, items_list) where:
    - broker_name: "cocos_capital", "invertir_online", "bull_market", or None
    - items_list: list of standardized transaction dicts

    If broker cannot be detected, returns (None, []).
    """
    broker = detect_broker(csv_bytes)

    if broker == "cocos_capital":
        items = parse_cocos_csv(csv_bytes)
    elif broker == "invertir_online":
        items = parse_invertir_online_csv(csv_bytes)
    elif broker == "bull_market":
        items = parse_bull_market_csv(csv_bytes)
    else:
        items = []

    return broker, items


def _detect_broker_from_rows(rows: list[list]) -> Optional[str]:
    """Auto-detect broker format from row headers.

    Returns one of: "cocos_capital", "invertir_online", "bull_market", or None.
    """
    if not rows:
        return None

    headers = _get_headers_lower(rows[0])

    if _is_cocos_account_statement(headers):
        return "cocos_capital"

    if "especie" in headers and "comision" in headers:
        return "cocos_capital"

    if "isin" in headers and "liquidacion" in headers:
        return "invertir_online"

    if "simbolo" in headers:
        return "bull_market"

    return None


def _rows_to_transactions(rows: list[list], broker: str) -> list[dict]:
    """Convert Excel rows to standardized transactions for given broker.

    Args:
        rows: List of lists (first row is headers)
        broker: One of "cocos_capital", "invertir_online", "bull_market"

    Returns:
        List of standardized transaction dicts
    """
    if not rows or broker not in ("cocos_capital", "invertir_online", "bull_market"):
        return []

    headers = _get_headers_lower(rows[0])

    if broker == "cocos_capital" and _is_cocos_account_statement(headers):
        return _parse_cocos_account_statement(rows)

    col_indices = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col_indices["fecha"] = i
        elif "tipo" in h:
            col_indices["tipo"] = i
        elif "especie" in h:
            col_indices["especie"] = i
        elif "simbolo" in h:
            col_indices["simbolo"] = i
        elif "isin" in h:
            col_indices["isin"] = i
        elif "cantidad" in h:
            col_indices["cantidad"] = i
        elif "precio" in h:
            col_indices["precio"] = i

    required = {"fecha", "cantidad", "precio"}
    if broker == "cocos_capital":
        required.add("tipo")
        required.add("especie")
    elif broker == "invertir_online":
        required.add("especie")
    elif broker == "bull_market":
        required.add("simbolo")

    if not required.issubset(col_indices.keys()):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col_indices.values()):
            continue

        date_val = row[col_indices["fecha"]]
        if isinstance(date_val, str):
            date = _parse_date(date_val)
        else:
            try:
                from datetime import datetime
                if isinstance(date_val, datetime):
                    date = date_val.strftime("%Y-%m-%d")
                else:
                    date = None
            except Exception:
                date = None

        if not date:
            continue

        if broker == "cocos_capital":
            tx_type_raw = str(row[col_indices["tipo"]]).strip()
            tx_type = "sell" if "venta" in tx_type_raw.lower() else "buy"
            ticker = str(row[col_indices["especie"]]).strip().upper()
        elif broker == "invertir_online":
            tx_type = "buy"
            ticker = str(row[col_indices["especie"]]).strip().upper()
        else:  # bull_market
            tx_type = "buy"
            ticker = str(row[col_indices["simbolo"]]).strip().upper()

        if not ticker:
            continue

        qty_val = row[col_indices["cantidad"]]
        quantity = _parse_amount(str(qty_val)) if not isinstance(qty_val, (int, float)) else float(qty_val)
        if quantity is None or quantity <= 0:
            continue

        price_val = row[col_indices["precio"]]
        price = _parse_amount(str(price_val)) if not isinstance(price_val, (int, float)) else float(price_val)
        if price is None or price <= 0:
            continue

        csv_reference = f"{broker}_{date}_{ticker}_{quantity}"

        items.append({
            "date": date,
            "tx_type": tx_type,
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": broker,
            "csv_reference": csv_reference,
        })

    return items


def _is_ppi_ledger_sheet(headers: list[str]) -> bool:
    """Detect a PPI (Portfolio Personal Inversiones) cash-ledger sheet.

    Columns: Fecha, Descripcion, Cantidad, Precio, Importe, Saldo, Moneda.
    PPI exports one ledger sheet per currency (e.g. 'Pesos', 'DolarCV7000 Ext.')
    plus an 'Instrumentos' summary sheet that lacks Importe/Saldo and is skipped.
    """
    return (
        "fecha" in headers
        and "cantidad" in headers
        and "precio" in headers
        and "importe" in headers
        and "saldo" in headers
    )


def _parse_ppi_ledger_rows(rows: list[list]) -> list[dict]:
    """Parse one PPI ledger sheet's raw rows (as returned by openpyxl values_only).

    Only rows whose Descripcion starts with 'COMPRA '/'VENTA ' are security trades;
    deposits, withdrawals, and cash dividends ('Ingreso de Fondos', 'Retiro de
    Fondos', 'Dividendo en efectivo') are skipped.
    """
    if not rows:
        return []

    headers = _get_headers_lower([str(h) if h is not None else "" for h in rows[0]])
    col = {}
    for i, h in enumerate(headers):
        if "fecha" in h:
            col["fecha"] = i
        elif "descrip" in h:
            col["descripcion"] = i
        elif "cantidad" in h:
            col["cantidad"] = i
        elif "precio" in h:
            col["precio"] = i

    required = {"fecha", "descripcion", "cantidad", "precio"}
    if not required.issubset(col.keys()):
        return []

    items = []
    for row in rows[1:]:
        if len(row) <= max(col.values()):
            continue

        descripcion = str(row[col["descripcion"]] or "").strip()
        lower_desc = descripcion.lower()
        if lower_desc.startswith("compra "):
            tx_type = "buy"
            ticker = descripcion[len("COMPRA "):].strip().upper()
        elif lower_desc.startswith("venta "):
            tx_type = "sell"
            ticker = descripcion[len("VENTA "):].strip().upper()
        else:
            continue

        if not ticker:
            continue

        date_val = row[col["fecha"]]
        if isinstance(date_val, str):
            date = _parse_date(date_val)
        else:
            from datetime import datetime
            date = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else None
        if not date:
            continue

        qty_val = row[col["cantidad"]]
        quantity = float(qty_val) if isinstance(qty_val, (int, float)) else _parse_amount(str(qty_val))
        if quantity is None or quantity == 0:
            continue
        quantity = abs(quantity)

        price_val = row[col["precio"]]
        price = float(price_val) if isinstance(price_val, (int, float)) else _parse_amount(str(price_val))
        if price is None or price <= 0:
            continue

        csv_reference = f"ppi_{date}_{ticker}_{quantity}"

        items.append({
            "date": date,
            "tx_type": tx_type,
            "ticker": ticker,
            "quantity": quantity,
            "price": price,
            "broker": "ppi",
            "csv_reference": csv_reference,
        })

    return items


def _parse_ppi_workbook(wb) -> list[dict]:
    """Parse all PPI ledger sheets in a workbook (one sheet per currency)."""
    items = []
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        headers = _get_headers_lower([str(h) if h is not None else "" for h in rows[0]])
        if _is_ppi_ledger_sheet(headers):
            items.extend(_parse_ppi_ledger_rows(rows))
    return items


def parse_xlsx(xlsx_bytes: bytes) -> Tuple[Optional[str], list[dict]]:
    """Parse Excel file (.xlsx) for investment transactions.

    Auto-detects broker format from headers. Single-sheet brokers (Cocos,
    Invertir Online, Bull Market) are checked first via the active sheet; if
    none match, falls back to PPI's multi-sheet (one ledger per currency) format.
    Returns tuple of (broker_name, items_list).
    """
    if not OPENPYXL_AVAILABLE:
        return None, []

    try:
        from io import BytesIO
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            row_list = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_list)

        if rows:
            broker = _detect_broker_from_rows(rows)
            if broker:
                items = _rows_to_transactions(rows, broker)
                return broker, items

        ppi_items = _parse_ppi_workbook(wb)
        if ppi_items:
            return "ppi", ppi_items

        return None, []

    except Exception:
        return None, []


def parse_file(file_bytes: bytes, filename: str) -> Tuple[Optional[str], list[dict]]:
    """Parse investment file (CSV or XLSX) by extension.

    Args:
        file_bytes: Raw file content
        filename: Original filename to detect format

    Returns:
        Tuple of (broker_name, items_list)
    """
    filename_lower = (filename or "").lower()

    if filename_lower.endswith(".xlsx"):
        return parse_xlsx(file_bytes)
    elif filename_lower.endswith(".csv"):
        return parse_csv(file_bytes)
    else:
        return None, []
